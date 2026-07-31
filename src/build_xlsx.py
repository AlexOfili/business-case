"""
Rebuild the live-formula unit-economics spreadsheet.

4 sheets:
  1. Assumptions     — blue inputs. Change a value → everything updates.
  2. PriceTiers      — profit & break-even at the 4 candidate prices.
  3. Sensitivity     — 3 stress scenarios (baseline, 2× support, 20% churn).
  4. ProfitCurve     — full price sweep £100→£350 for the chart.

Run:
    python3 build_xlsx.py
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(os.path.dirname(HERE), "unit_economics.xlsx")

BLUE   = Font(color="0000FF", bold=True)            # inputs
BOLD   = Font(bold=True)
HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SOFT   = PatternFill("solid", fgColor="E8EEF7")
WARN   = PatternFill("solid", fgColor="FFE8E8")
THIN   = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
def build():
    wb = Workbook()

    # ---- Assumptions sheet ----
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "Riverside Malls — Smart Trolley Pilot"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "Inputs (blue) drive every formula on the other sheets."
    ws["A2"].font = Font(italic=True, color="555555")

    headers = ["Assumption", "Value", "Note"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, 3)

    inputs = [
        ("Variable cost per trolley (£/mo)", 40,    "Maintenance, support, telematics"),
        ("Fixed cost per month (£)",         3000,  "Platform, depot, on-call staff"),
        ("Number of trolleys in pilot",      25,    "Phase-1 fleet size"),
        ("Price tier 1 — floor (£)",         150,   "Penetration price"),
        ("Price tier 2 — base (£)",          200,   "Headline pilot price"),
        ("Price tier 3 — premium (£)",       250,   "Stretch price"),
        ("Price tier 4 — aspirational (£)",  300,   "Upside price"),
        ("Churn rate (stretch, monthly)",    0.20,  "Trolleys needing re-deployment"),
        ("Re-deployment cost per churned trolley (£)", 120,
            "Pickup + refurb + logistics"),
        ("Support cost multiplier (stretch)", 2,    "2× = stressed scenario"),
    ]
    for i, (label, val, note) in enumerate(inputs, start=5):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=val)
        c.font = BLUE
        ws.cell(row=i, column=3, value=note)
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = BORDER
    autosize(ws, [42, 12, 40])

    # Named-cell-style absolute refs (always cross-sheet to Assumptions)
    V  = "Assumptions!$B$5"; F_ = "Assumptions!$B$6"; T_ = "Assumptions!$B$7"
    P1 = "Assumptions!$B$8"; P2 = "Assumptions!$B$9"
    P3 = "Assumptions!$B$10"; P4 = "Assumptions!$B$11"
    CH = "Assumptions!$B$12"; RC = "Assumptions!$B$13"
    SM = "Assumptions!$B$14"

    # ---- PriceTiers ----
    pt = wb.create_sheet("PriceTiers")
    pt["A1"] = "Profit & Break-even at the 4 Candidate Prices"
    pt["A1"].font = Font(size=13, bold=True)
    pt_headers = ["Price tier", "Price (£)", "Contribution (£)",
                  "Monthly profit (£)", "Break-even trolleys"]
    for c, h in enumerate(pt_headers, 1):
        pt.cell(row=3, column=c, value=h)
    style_header(pt, 3, 5)

    tiers = [
        ("Floor (£150)",         P1),
        ("Base (£200)",          P2),
        ("Premium (£250)",       P3),
        ("Aspirational (£300)",  P4),
    ]
    for i, (label, pref) in enumerate(tiers, start=4):
        pt.cell(row=i, column=1, value=label)
        pt.cell(row=i, column=2, value=f"={pref}")
        pt.cell(row=i, column=3, value=f"=B{i}-{V}")
        pt.cell(row=i, column=4, value=f"={T_}*C{i}-{F_}")
        pt.cell(row=i, column=5, value=f"={F_}/C{i}")
        for col in range(1, 6):
            pt.cell(row=i, column=col).border = BORDER
        # Highlight base tier
        if label.startswith("Base"):
            for col in range(1, 6):
                pt.cell(row=i, column=col).fill = SOFT
                pt.cell(row=i, column=col).font = BOLD

    pt["A9"] = "Recommendation:"
    pt["A9"].font = BOLD
    pt["B9"] = ("Charge £200/trolley/month. £1,000/month profit, "
                "6-trolley margin of safety above the 18.75 break-even.")
    pt.merge_cells("B9:E9")
    autosize(pt, [22, 12, 18, 18, 22])

    # ---- Sensitivity ----
    se = wb.create_sheet("Sensitivity")
    se["A1"] = "Sensitivity — How much margin can we lose before the pilot breaks?"
    se["A1"].font = Font(size=13, bold=True)
    se_headers = ["Scenario", "Variable cost (£)", "Fixed cost (£)",
                  "Effective trolleys", "Monthly profit (£) @ £200"]
    for c, h in enumerate(se_headers, 1):
        se.cell(row=3, column=c, value=h)
    style_header(se, 3, 5)

    se_rows = [
        ("Baseline",
         f"={V}", f"={F_}", f"={T_}",
         f"=D4*({P2}-B4)-C4"),
        ("Support cost doubles",
         f"={V}*{SM}", f"={F_}", f"={T_}",
         f"=D5*({P2}-B5)-C5"),
        ("20% monthly churn (re-deployment cost)",
         f"={V}", f"={F_}+{CH}*{T_}*{RC}", f"={T_}*(1-{CH})",
         f"=D6*({P2}-B6)-C6"),
    ]
    for i, (label, vb, fc, et, pf) in enumerate(se_rows, start=4):
        se.cell(row=i, column=1, value=label)
        se.cell(row=i, column=2, value=vb)
        se.cell(row=i, column=3, value=fc)
        se.cell(row=i, column=4, value=et)
        se.cell(row=i, column=5, value=pf)
        for col in range(1, 6):
            se.cell(row=i, column=col).border = BORDER
        if "churn" in label.lower():
            for col in range(1, 6):
                se.cell(row=i, column=col).fill = WARN

    se["A8"] = ("Read: at £200, the pilot survives 2× support cost exactly at "
                "break-even. It does NOT survive 20% monthly churn.")
    se["A8"].font = Font(italic=True)
    se.merge_cells("A8:E8")
    autosize(se, [36, 18, 14, 18, 28])

    # ---- ProfitCurve (full sweep) ----
    pc = wb.create_sheet("ProfitCurve")
    pc["A1"] = "Full Price Sweep — for profit-vs-price chart"
    pc["A1"].font = Font(size=13, bold=True)
    for c, h in enumerate(["Price (£)", "Monthly profit (£)",
                           "Break-even trolleys"], 1):
        pc.cell(row=3, column=c, value=h)
    style_header(pc, 3, 3)

    for i, price in enumerate(range(100, 351, 10), start=4):
        pc.cell(row=i, column=1, value=price)
        pc.cell(row=i, column=2, value=f"={T_}*(A{i}-{V})-{F_}")
        pc.cell(row=i, column=3, value=f"={F_}/(A{i}-{V})")
        for col in range(1, 4):
            pc.cell(row=i, column=col).border = BORDER
    autosize(pc, [14, 20, 22])

    # Save
    wb.save(OUT)
    print(f"  wrote {OUT}")
    return OUT


if __name__ == "__main__":
    print("Rebuilding unit_economics.xlsx...")
    build()
    print("Done.")
